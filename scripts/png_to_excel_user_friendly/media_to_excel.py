#!/usr/bin/env python3
"""
Media to Excel Converter
=======================

This script converts images and PDFs in a specified folder into an Excel file with thumbnails and metadata.

The script will guide you through:
1. Selecting the input folder containing your media files
2. Choosing where to save the output
3. Naming your output Excel file

Requirements:
- Python 3.9+
- Pillow
- openpyxl
- pdf2image (for PDF support)

HOW TO MODIFY THIS SCRIPT:
-------------------------
1. To add new columns:
   - Find the 'headers' list below (around line 80)
   - Add your new column name to the list
   - Add the corresponding data in the loop below (around line 100)
   - Example: To add a "File Size" column:
     * Add "File Size" to headers list
     * Add: ws.cell(row=row, column=9, value=os.path.getsize(path))

2. To change thumbnail size:
   - Find the line with 'thumb.thumbnail((256, 256)' (around line 95)
   - Change the numbers to your desired size (width, height)

3. To change DPI for inch calculations:
   - Find the line with 'w_in = w / 300' (around line 90)
   - Change 300 to your desired DPI

4. To change Excel column widths:
   - Find the section with 'ws.column_dimensions' (around line 80)
   - Adjust the width values (currently set to 20)
"""
import os
import sys
import subprocess
import importlib.util
from pathlib import Path

# List of required packages - don't modify unless you know what you're doing
REQUIRED_PACKAGES = ["Pillow", "openpyxl", "pdf2image"]

# Check and install missing packages
for pkg in REQUIRED_PACKAGES:
    if importlib.util.find_spec(pkg) is None:
        print(f"Installing missing package: {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

try:
    from PIL import Image
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from pdf2image import convert_from_path
except ImportError as e:
    print(f"Failed to import required modules: {e}")
    sys.exit(1)

# Supported file extensions
SUPPORTED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.tif', '.tiff', '.pdf'}

def get_image_from_file(file_path):
    """Get image from file, handling both images and PDFs."""
    ext = Path(file_path).suffix.lower()
    if ext == '.pdf':
        # Convert first page of PDF to image
        images = convert_from_path(file_path, first_page=1, last_page=1)
        if images:
            return images[0]
        return None
    else:
        # Open image file
        return Image.open(file_path)

def get_user_input(prompt, default=None):
    """Get user input with optional default value."""
    if default:
        user_input = input(f"{prompt} [{default}]: ").strip()
        return user_input if user_input else default
    return input(f"{prompt}: ").strip()

def main():
    print("\n=== Media to Excel Converter ===\n")
    
    # Step 1: Get input directory
    while True:
        input_dir = get_user_input("Enter the folder path containing your media files")
        if os.path.isdir(input_dir):
            break
        print(f"Error: '{input_dir}' is not a valid directory. Please try again.")

    # Step 2: Get output directory
    while True:
        output_dir = get_user_input("Enter the folder path where you want to save the results")
        try:
            os.makedirs(output_dir, exist_ok=True)
            # Create thumbnails subfolder
            thumbnails_dir = os.path.join(output_dir, "thumbnails")
            os.makedirs(thumbnails_dir, exist_ok=True)
            break
        except Exception as e:
            print(f"Error creating output directory: {e}. Please try again.")

    # Step 3: Get Excel filename
    default_name = "media_info"
    excel_name = get_user_input("Enter a name for your Excel file (without .xlsx)", default_name)
    if not excel_name.endswith('.xlsx'):
        excel_name += '.xlsx'

    # Get list of supported files
    media_files = [f for f in os.listdir(input_dir) 
                  if Path(f).suffix.lower() in SUPPORTED_EXTENSIONS]
    
    if not media_files:
        print(f"\nNo supported files found in '{input_dir}'")
        print(f"Supported extensions: {', '.join(SUPPORTED_EXTENSIONS)}")
        sys.exit(0)

    print(f"\nProcessing {len(media_files)} files...")

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Info"

    # Define column headers - Add or modify columns here
    headers = [
        "Image",           # Column A: Thumbnail preview
        "Filename",        # Column B: Original filename
        "Width (px)",      # Column C: Width in pixels
        "Height (px)",     # Column D: Height in pixels
        "Width (in)",      # Column E: Width in inches (at 300 DPI)
        "Height (in)",     # Column F: Height in inches (at 300 DPI)
        "Mode",           # Column G: Color mode (RGB, RGBA, etc.)
        "Format"          # Column H: Original file format
    ]
    ws.append(headers)

    # Set column widths - Adjust these numbers to change column widths
    ws.column_dimensions[get_column_letter(1)].width = 20  # Image column
    for i in range(2, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Process each file
    row = 2  # Start from row 2 (row 1 has headers)
    for fname in media_files:
        path = os.path.join(input_dir, fname)
        try:
            img = get_image_from_file(path)
            if img is None:
                print(f"Skipping {fname}: Could not process file")
                continue
        except Exception as e:
            print(f"Skipping {fname}: {e}")
            continue

        # Get image dimensions and convert to inches (at 300 DPI)
        w, h = img.size
        w_in = w / 300  # Change 300 to adjust DPI
        h_in = h / 300  # Change 300 to adjust DPI
        mode = img.mode
        fmt = Path(fname).suffix[1:].upper()  # Get original file extension

        # Create thumbnail
        thumb = img.copy()
        thumb.thumbnail((256, 256), Image.LANCZOS)  # Increased thumbnail size
        thumb_path = os.path.join(thumbnails_dir, f"thumb_{Path(fname).stem}.png")  # Save as PNG
        thumb.save(thumb_path, "PNG")

        # Add data to Excel
        xl_img = XLImage(thumb_path)
        ws.row_dimensions[row].height = 150  # Increased row height for larger thumbnails
        ws.add_image(xl_img, f"A{row}")  # Add thumbnail to column A
        ws.cell(row=row, column=2, value=fname)  # Add filename to column B
        ws.cell(row=row, column=3, value=w)      # Add width to column C
        ws.cell(row=row, column=4, value=h)      # Add height to column D
        ws.cell(row=row, column=5, value=round(w_in, 2))  # Add width in inches to column E
        ws.cell(row=row, column=6, value=round(h_in, 2))  # Add height in inches to column F
        ws.cell(row=row, column=7, value=mode)   # Add color mode to column G
        ws.cell(row=row, column=8, value=fmt)    # Add original format to column H
        row += 1

    # Save Excel file
    out_path = os.path.join(output_dir, excel_name)
    wb.save(out_path)
    print(f"\nDone! Excel sheet saved to:\n{out_path}")
    print(f"Thumbnails saved to:\n{thumbnails_dir}")

if __name__ == "__main__":
    main() 