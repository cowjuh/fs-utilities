#!/usr/bin/env python3
"""
Media to Excel Converter
=======================

This script converts images and PDFs in a specified folder into an Excel file with thumbnails and metadata.

Usage:
    python3 media_to_excel.py -i /path/to/input/folder -o /path/to/output/folder [-n output_name]

Options:
    -i, --input     Path to the folder containing media files (required)
    -o, --output    Path to the output folder (required)
    -n, --name      Name for the output Excel file (default: media_info)

Requirements:
- Python 3.9+
- Pillow
- openpyxl
- pdf2image (for PDF support)

HOW TO MODIFY THIS SCRIPT:
-------------------------
1. To add new columns:
   - Find the 'headers' list below (around line 80)
   - Add your new column name to the list
   - Add the corresponding data in the loop below (around line 100)
   - Example: To add a "File Size" column:
     * Add "File Size" to headers list
     * Add: ws.cell(row=row, column=9, value=os.path.getsize(path))

2. To change thumbnail size:
   - Find the line with 'thumb.thumbnail((256, 256)' (around line 95)
   - Change the numbers to your desired size (width, height)

3. To change DPI for inch calculations:
   - Find the line with 'w_in = w / 300' (around line 90)
   - Change 300 to your desired DPI

4. To change Excel column widths:
   - Find the section with 'ws.column_dimensions' (around line 80)
   - Adjust the width values (currently set to 20)
"""
import os
import sys
import subprocess
import importlib.util
import argparse
from pathlib import Path
import io
from datetime import datetime
import openpyxl.styles

# List of required packages - don't modify unless you know what you're doing
REQUIRED_PACKAGES = ["Pillow", "openpyxl", "pdf2image"]

# Check and install missing packages
for pkg in REQUIRED_PACKAGES:
    if importlib.util.find_spec(pkg) is None:
        print(f"Installing missing package: {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

try:
    from PIL import Image
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from pdf2image import convert_from_path
    from PIL import ImageCms
except ImportError as e:
    print(f"Failed to import required modules: {e}")
    sys.exit(1)

# Supported file extensions
SUPPORTED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.tif', '.tiff', '.pdf'}

def get_image_from_file(file_path):
    """Get image from file, handling both images and PDFs."""
    ext = Path(file_path).suffix.lower()
    if ext == '.pdf':
        # Convert first page of PDF to image
        images = convert_from_path(file_path, first_page=1, last_page=1)
        if images:
            return images[0]
        return None
    else:
        # Open image file
        return Image.open(file_path)

def main():
    # Set up command line arguments
    parser = argparse.ArgumentParser(description='Convert media files in a folder to an Excel spreadsheet with metadata.')
    parser.add_argument('-i', '--input', required=True, help='Path to the folder containing media files')
    parser.add_argument('-o', '--output', required=True, help='Path to the output folder')
    parser.add_argument('-n', '--name', default='media_info', help='Name for the output Excel file (without .xlsx)')
    args = parser.parse_args()

    # Get input and output directories
    input_dir = args.input
    output_dir = args.output
    excel_name = args.name

    # Generate timestamp for unique naming
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Check if input directory exists
    if not os.path.isdir(input_dir):
        print(f"Error: Input directory '{input_dir}' is not a valid directory")
        sys.exit(1)

    # Create output directory and thumbnails subfolder
    try:
        os.makedirs(output_dir, exist_ok=True)
        thumbnails_dir = os.path.join(output_dir, f"thumbnails_{timestamp}")
        os.makedirs(thumbnails_dir, exist_ok=True)
    except Exception as e:
        print(f"Error creating output directory: {e}")
        sys.exit(1)

    # Add .xlsx extension if not present and append timestamp
    if not excel_name.endswith('.xlsx'):
        excel_name = f"{excel_name}_{timestamp}.xlsx"
    else:
        excel_name = f"{excel_name[:-5]}_{timestamp}.xlsx"

    # Get list of supported files
    media_files = [f for f in os.listdir(input_dir) 
                  if Path(f).suffix.lower() in SUPPORTED_EXTENSIONS]
    
    if not media_files:
        print(f"No supported files found in '{input_dir}'")
        print(f"Supported extensions: {', '.join(SUPPORTED_EXTENSIONS)}")
        sys.exit(0)

    print(f"Processing {len(media_files)} files...")

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Info"

    # Define column headers - Add or modify columns here
    headers = [
        "Image",           # Column A: Thumbnail preview
        "Filename",        # Column B: Original filename
        "Width (px)",      # Column C: Width in pixels
        "Height (px)",     # Column D: Height in pixels
        "Width (in) @ 300 DPI",      # Column E: Width in inches (at 300 DPI)
        "Height (in) @ 300 DPI",     # Column F: Height in inches (at 300 DPI)
        "Print Dimensions @ 300DPI",  # Column G: Combined dimensions in inches
        "Mode",           # Column H: Color mode (RGB, RGBA, etc.)
        "Format",         # Column I: Original file format
        "Color Profile"   # Column J: Color profile information
    ]
    ws.append(headers)

    # Set column widths - Adjust these numbers to change column widths
    ws.column_dimensions[get_column_letter(1)].width = 35  # Image column - increased for 256px thumbnails
    ws.column_dimensions[get_column_letter(2)].width = 50  # Filename column - increased for long filenames
    for i in range(3, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Process each file
    row = 2  # Start from row 2 (row 1 has headers)
    for fname in media_files:
        path = os.path.join(input_dir, fname)
        try:
            img = get_image_from_file(path)
            if img is None:
                print(f"Skipping {fname}: Could not process file")
                continue
        except Exception as e:
            print(f"Skipping {fname}: {e}")
            continue

        # Get image dimensions and convert to inches (at 300 DPI)
        w, h = img.size
        w_in = w / 300  # Change 300 to adjust DPI
        h_in = h / 300  # Change 300 to adjust DPI
        mode = img.mode
        fmt = Path(fname).suffix[1:].upper()  # Get original file extension
        
        # Get color profile information
        try:
            icc_profile = img.info.get('icc_profile', None)
            if icc_profile:
                try:
                    prf = ImageCms.ImageCmsProfile(io.BytesIO(icc_profile))
                    color_profile = ImageCms.getProfileDescription(prf)
                except Exception:
                    color_profile = "Unknown Profile"
            else:
                color_profile = "Untagged"
            # Sanitize color_profile to remove illegal characters for Excel
            import string
            color_profile = ''.join(c for c in color_profile if c in string.printable and c not in '\x00\x01\x02\x03\x04\x05\x06\x07\x08\x0b\x0c\x0e\x0f\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1a\x1b\x1c\x1d\x1e\x1f')
        except Exception:
            color_profile = "Profile Info Unavailable"

        # Create thumbnail
        thumb = img.copy()
        thumb.thumbnail((256, 256), Image.LANCZOS)  # Increased thumbnail size
        thumb_path = os.path.join(thumbnails_dir, f"thumb_{Path(fname).stem}.png")  # Save as PNG
        thumb.save(thumb_path, "PNG")

        # Add data to Excel
        xl_img = XLImage(thumb_path)
        ws.row_dimensions[row].height = 192  # Increased row height for better thumbnail display
        ws.add_image(xl_img, f"A{row}")
        
        # Set cell values and alignment
        cells = [
            (2, fname),              # Filename
            (3, w),                  # Width in pixels
            (4, h),                  # Height in pixels
            (5, round(w_in, 2)),     # Width in inches
            (6, round(h_in, 2)),     # Height in inches
            (7, f'{w_in:.2f}" x {h_in:.2f}"'),  # Combined dimensions with fixed decimal places
            (8, mode),               # Color mode
            (9, fmt),                # Original format
            (10, color_profile)      # Color profile info
        ]
        
        for col, value in cells:
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        
        row += 1

    # Auto-adjust column widths based on content
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Skip auto-adjusting the image column (A)
        if column_letter == 'A':
            continue
            
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Add some padding to the width
        adjusted_width = (max_length + 2)
        # Set a minimum width of 10 and maximum of 100
        adjusted_width = max(10, min(adjusted_width, 100))
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save Excel file
    out_path = os.path.join(output_dir, excel_name)
    wb.save(out_path)
    print(f"\nDone! Excel sheet saved to:\n{out_path}")
    
    # Clean up thumbnails folder
    try:
        import shutil
        shutil.rmtree(thumbnails_dir)
        print("Thumbnails folder cleaned up successfully.")
    except Exception as e:
        print(f"Warning: Could not clean up thumbnails folder: {e}")

if __name__ == "__main__":
    main() 