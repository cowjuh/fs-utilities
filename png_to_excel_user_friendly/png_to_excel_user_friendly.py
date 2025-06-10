#!/usr/bin/env python3
"""
PNG to Excel Converter
======================

This script converts all PNG images in a selected folder into an Excel file with thumbnails and metadata.

How to use:
1. Double-click or run this script. It will prompt you to select a folder.
2. The script will check and install any missing dependencies (Pillow, openpyxl, tkinter).
3. The resulting Excel file will be saved in a new folder next to your images.

If you have issues, ensure you have Python 3.9+ and Homebrew installed.
"""
import os
import sys
import subprocess
import importlib.util

# List of required packages
REQUIRED_PACKAGES = ["Pillow", "openpyxl"]

# Check and install missing packages
for pkg in REQUIRED_PACKAGES:
    if importlib.util.find_spec(pkg) is None:
        print(f"Installing missing package: {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

try:
    from PIL import Image
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    import tkinter as tk
    from tkinter import filedialog, messagebox
except ImportError as e:
    print(f"Failed to import required modules: {e}")
    sys.exit(1)

def select_folder():
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title="Select Folder Containing PNG Files")
    if not folder:
        messagebox.showinfo("No Folder Selected", "No folder was selected. Exiting.")
        sys.exit(0)
    return folder

def main():
    SRC_DIR = select_folder()
    BASE_DIR = os.path.dirname(SRC_DIR)
    FOLDER_NAME = os.path.basename(SRC_DIR)
    DEST_DIR = os.path.join(BASE_DIR, FOLDER_NAME + " EXCEL")
    os.makedirs(DEST_DIR, exist_ok=True)

    png_files = [f for f in os.listdir(SRC_DIR) if f.lower().endswith('.png')]
    if not png_files:
        messagebox.showinfo("No PNG Files", "No PNG files found in the selected folder.")
        sys.exit(0)

    wb = Workbook()
    ws = wb.active
    ws.title = "PNG Info"
    headers = ["Image", "Filename", "Width (px)", "Height (px)", "Width (in)", "Height (in)", "Mode", "Format"]
    ws.append(headers)
    ws.column_dimensions[get_column_letter(1)].width = 20
    for i in range(2, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row = 2
    for fname in png_files:
        path = os.path.join(SRC_DIR, fname)
        try:
            img = Image.open(path)
        except Exception as e:
            print(f"Skipping {fname}: {e}")
            continue
        w, h = img.size
        w_in = w / 300
        h_in = h / 300
        mode = img.mode
        fmt = img.format or 'PNG'
        thumb = img.copy()
        thumb.thumbnail((128, 128), Image.LANCZOS)
        thumb_path = os.path.join(DEST_DIR, f"thumb_{fname}")
        thumb.save(thumb_path)
        xl_img = XLImage(thumb_path)
        ws.row_dimensions[row].height = 100
        ws.add_image(xl_img, f"A{row}")
        ws.cell(row=row, column=2, value=fname)
        ws.cell(row=row, column=3, value=w)
        ws.cell(row=row, column=4, value=h)
        ws.cell(row=row, column=5, value=round(w_in, 2))
        ws.cell(row=row, column=6, value=round(h_in, 2))
        ws.cell(row=row, column=7, value=mode)
        ws.cell(row=row, column=8, value=fmt)
        row += 1

    out_path = os.path.join(DEST_DIR, 'png_info.xlsx')
    wb.save(out_path)
    messagebox.showinfo("Done", f"Excel sheet saved to:\n{out_path}")

if __name__ == "__main__":
    main() 